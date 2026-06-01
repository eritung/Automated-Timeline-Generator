
import io
import re
import uuid
from datetime import date, datetime, timedelta
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Side
from openpyxl.cell.cell import MergedCell


st.set_page_config(page_title="製作時程排程工具", page_icon="📅", layout="wide", initial_sidebar_state="collapsed")

# =========================
# 基本設定
# =========================
MODE_OPTIONS = ["製作日推進", "上線日回推", "同時指定開始與上線日期"]
MODE_MAP = {
    "製作日推進": "forward",
    "上線日回推": "backward",
    "同時指定開始與上線日期": "double",
}
OWNER_OPTIONS = ["Ad2", "客戶", "Ad2＋客戶"]
OWNER_ALIASES = {
    "Ad2+客戶": "Ad2＋客戶",
    "Ad2＋客戶": "Ad2＋客戶",
    "AD2+客戶": "Ad2＋客戶",
    "AD2＋客戶": "Ad2＋客戶",
    "Ad2&客戶": "Ad2＋客戶",
    "Ad2＆客戶": "Ad2＋客戶",
    "Ad2/客戶": "Ad2＋客戶",
    "AD2": "Ad2",
    "ad2": "Ad2",
}

def normalize_owner(owner: str, fallback_text: str = "") -> str:
    """統一 Action By 顯示；支援 Ad2＋客戶這類共同執行項目。"""
    owner = str(owner or "").strip()
    owner = OWNER_ALIASES.get(owner, owner)
    if owner in OWNER_OPTIONS:
        return owner
    return "客戶" if "客戶" in str(fallback_text) else "Ad2"


def format_day_value(days):
    """工作天數顯示用：1.0 -> 1、0.5 -> 0.5。"""
    try:
        value = float(days)
    except (TypeError, ValueError):
        return days
    return int(value) if value.is_integer() else value

DEFAULT_HALF_DAY_LABEL = "1300"

DEFAULT_TASKS = [
    {"id": "task_1", "顯示": True, "任務名稱": "提供素材", "Action By": "客戶", "工作天數": 1.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_2", "顯示": True, "任務名稱": "視覺製作", "Action By": "Ad2", "工作天數": 3.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_3", "顯示": True, "任務名稱": "客戶確認", "Action By": "客戶", "工作天數": 1.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_4", "顯示": True, "任務名稱": "視覺調整", "Action By": "Ad2", "工作天數": 2.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_5", "顯示": True, "任務名稱": "客戶確認", "Action By": "客戶", "工作天數": 1.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_6", "顯示": True, "任務名稱": "廣告進稿", "Action By": "Ad2", "工作天數": 1.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": False, "粗下框線": False},
    {"id": "task_7", "顯示": True, "任務名稱": "廣告上線", "Action By": "Ad2", "工作天數": 1.0, "半天標註": DEFAULT_HALF_DAY_LABEL, "上線日": True, "粗下框線": False},
]

DEFAULT_BATCH_TASKS_TEXT = """提供素材 客戶 1天
視覺製作 Ad2 3天
客戶確認 客戶 1天
視覺調整 Ad2 2天
客戶確認 客戶 1天
廣告上線 Ad2 1天 上線"""

WEBSITE_BATCH_TASKS_TEXT = """網站架構 2天
客戶確認 1
網站架構調整 2
客戶確認 1
視覺製作 5
客戶確認 1
網頁視覺調整 3
客戶確認 1
網頁切版 4
動態程式 4
客戶確認 1
功能測試 2
客戶確認 1
網頁打包與測試 2
客戶確認 1
廣告進稿 1
網站上線 1 上線"""

BATCH_TEMPLATE_OPTIONS = ["一般製作時程", "網頁製作時程"]
BATCH_TEMPLATE_MAP = {
    "一般製作時程": DEFAULT_BATCH_TASKS_TEXT,
    "網頁製作時程": WEBSITE_BATCH_TASKS_TEXT,
}

# 載入批次範本時，僅更新批次輸入內容，不應影響專案設定區的日期與排程方式。
PROJECT_SETTING_KEYS = [
    "project_name",
    "mode_display",
    "start_date_value",
    "launch_date_value",
    "collapse_threshold",
]

DEFAULT_HOLIDAYS = {
    '2025-12-25': '行憲紀念日',
    '2026-01-01': '元旦',
    '2026-01-21': 'Ad2尾牙',
    '2026-02-15': '春節連假',
    '2026-02-16': '春節連假',
    '2026-02-17': '春節連假',
    '2026-02-18': '春節連假',
    '2026-02-19': '春節連假',
    '2026-02-20': '春節連假',
    '2026-02-27': '二二八連假',
    '2026-02-28': '二二八連假',
    '2026-04-03': '清明連假',
    '2026-04-04': '清明連假',
    '2026-04-05': '清明連假',
    '2026-04-06': '清明連假',
    '2026-05-01': '勞動節放假',
    '2026-06-19': '端午節連假',
    '2026-09-25': '國定連假',
    '2026-09-28': '國定連假',
    '2026-10-09': '國慶日連假',
    '2026-10-10': '國慶日連假',
    '2026-10-25': '光復節連假',
    '2026-10-26': '光復節連假',
    '2026-12-25': '行憲紀念日連假',
}

# Excel colors
EXCEL_COLOR_CLIENT_BAR = '#EA9B56'
EXCEL_COLOR_AD2_BAR = '#4BACC6'
EXCEL_COLOR_LAUNCH_BAR = '#FF0000'
EXCEL_COLOR_PREP_BAR = '#92D050'
EXCEL_COLOR_WEEKEND = '#D9D9D9'
EXCEL_COLOR_HOLIDAY_TEXT = '#595959'
# 月份色固定對照：依當月代表節慶／節日延伸色票，盡量避免月份顏色重複。
MONTH_COLORS = {
    1: '#F4CCCC',   # 1月｜新年／春節前夕｜喜氣淡紅
    2: '#FCE4D6',   # 2月｜元宵節｜燈籠暖橘
    3: '#FCE4EC',   # 3月｜櫻花季／女神節｜櫻花粉
    4: '#D9EAD3',   # 4月｜清明／植栽感｜艾草綠
    5: '#EADCF8',   # 5月｜母親節｜康乃馨粉紫
    6: '#D9EAF7',   # 6月｜端午／水面感｜龍舟水藍
    7: '#FFF2CC',   # 7月｜暑假｜芒果奶黃
    8: '#D0E0E3',   # 8月｜父親節／夏末海風｜沉靜藍綠
    9: '#EADBC8',   # 9月｜中秋節｜月餅米杏
    10: '#FCE5CD',  # 10月｜萬聖節｜南瓜杏橘
    11: '#D9D2E9',  # 11月｜感恩節／年末採購｜葡萄紫
    12: '#DDEED6',  # 12月｜聖誕節｜聖誕柔綠
}

def get_month_color(month):
    return MONTH_COLORS.get(int(month), '#E8DFD0')

# UI colors — 液態玻璃風格
UI_PRIMARY = "#5B8FD4"        # 玻璃藍
UI_PRIMARY_HOVER = "#4A7BC4"
UI_BORDER = "rgba(180,200,230,0.45)"
UI_MUTED = "#8899BB"
UI_AD2 = "#5BA4D4"            # 玻璃藍
UI_CLIENT = "#E8956A"         # 珊瑚橙
UI_LAUNCH = "#E05C7A"         # 玫瑰紅
UI_PREP = "#6EC4A0"           # 薄荷綠

st.markdown(f"""
<style>
/*
 * 字體策略：Noto Sans TC（思源黑體，Google Fonts 跨平台）
 * Mac 備援：PingFang TC / Helvetica Neue
 * Windows 備援：Microsoft JhengHei
 * 通用備援：sans-serif
 */
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=block');
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@300;400;500;600&display=swap');

/* ══════════════════════════════════════
   全域基底 — 霧玻璃 / Liquid Glass 風格
   背景主色：純淨淺白，帶極淡的藍紫漸層光暈
   ══════════════════════════════════════ */
html, body, [data-testid="stAppViewContainer"],
[data-testid="stAppViewContainer"] * {{
  font-family: 'Noto Sans TC', 'PingFang TC', 'Microsoft JhengHei', 'Helvetica Neue', sans-serif !important;
  -webkit-font-smoothing: antialiased;
  line-height: 1.85 !important;
  letter-spacing: 0.04em !important;
}}

/* 主背景：淺白 + 彩色光暈裝飾 */
html, body {{
  background: #F7F8FC !important;
  color: #1A1D2E;
}}
[data-testid="stAppViewContainer"] {{
  background:
    radial-gradient(ellipse 60% 40% at 15% 20%, rgba(120,160,255,0.10) 0%, transparent 70%),
    radial-gradient(ellipse 50% 50% at 85% 75%, rgba(180,120,255,0.08) 0%, transparent 65%),
    radial-gradient(ellipse 40% 35% at 55% 90%, rgba(80,210,180,0.07) 0%, transparent 60%),
    #F7F8FC !important;
  color: #1A1D2E;
}}
[data-testid="stAppViewBlock"] {{
  background: transparent !important;
}}

/* ── Sidebar 展開按鈕 ── */
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"],
button[title="Open sidebar"],
button[title="Close sidebar"],
button[aria-label="Open sidebar"],
button[aria-label="Close sidebar"] {{
  width: 2.35rem !important;
  height: 2.35rem !important;
  min-width: 2.35rem !important;
  min-height: 2.35rem !important;
  border: 1px solid rgba(160,185,240,0.5) !important;
  border-radius: 999px !important;
  background: rgba(255,255,255,0.72) !important;
  backdrop-filter: blur(12px) saturate(1.4) !important;
  -webkit-backdrop-filter: blur(12px) saturate(1.4) !important;
  box-shadow: 0 2px 10px rgba(91,143,212,0.12), 0 1px 3px rgba(0,0,0,0.06) !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  overflow: hidden !important;
}}
[data-testid="collapsedControl"]:hover,
[data-testid="stSidebarCollapsedControl"]:hover,
button[title="Open sidebar"]:hover,
button[title="Close sidebar"]:hover,
button[aria-label="Open sidebar"]:hover,
button[aria-label="Close sidebar"]:hover {{
  background: rgba(235,242,255,0.88) !important;
  border-color: rgba(91,143,212,0.5) !important;
}}
[data-testid="collapsedControl"] *,
[data-testid="stSidebarCollapsedControl"] *,
button[title="Open sidebar"] *,
button[title="Close sidebar"] *,
button[aria-label="Open sidebar"] *,
button[aria-label="Close sidebar"] * {{
  font-family: 'Material Symbols Rounded' !important;
  font-weight: normal !important;
  font-style: normal !important;
  font-size: 24px !important;
  line-height: 1 !important;
  letter-spacing: normal !important;
  text-transform: none !important;
  display: inline-flex !important;
  align-items: center !important;
  justify-content: center !important;
  color: {UI_PRIMARY} !important;
  font-feature-settings: 'liga' !important;
  -webkit-font-feature-settings: 'liga' !important;
  font-variation-settings: 'FILL' 0, 'wght' 400, 'GRAD' 0, 'opsz' 24 !important;
  white-space: nowrap !important;
}}
span[class*="material"],
[data-testid="stIconMaterial"] {{
  font-family: 'Material Symbols Rounded' !important;
  letter-spacing: normal !important;
  line-height: 1 !important;
  font-feature-settings: 'liga' !important;
  -webkit-font-feature-settings: 'liga' !important;
}}
svg, svg * {{
  font-family: inherit;
  line-height: 1 !important;
  letter-spacing: normal !important;
}}

/* ── 下拉選單文字對齊修正 ── */
[data-baseweb="select"] > div,
[data-baseweb="select"] > div > div,
[data-baseweb="select"] [role="listbox"],
[data-baseweb="select"] [role="option"],
[data-baseweb="select"] span,
[data-baseweb="select"] div {{
  line-height: 1.4 !important;
  display: flex !important;
  align-items: center !important;
}}
[data-baseweb="select"] > div {{
  min-height: 2.6rem !important;
  padding-top: 0 !important;
  padding-bottom: 0 !important;
}}

.block-container {{
  max-width: 1480px;
  padding-top: 2.2rem !important;
  padding-bottom: 4rem !important;
}}

/* ── Sidebar 玻璃風格 ── */
[data-testid="stSidebar"] {{
  background: rgba(255,255,255,0.75) !important;
  backdrop-filter: blur(20px) saturate(1.5) !important;
  -webkit-backdrop-filter: blur(20px) saturate(1.5) !important;
  border-right: 1px solid rgba(160,185,240,0.3) !important;
}}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] textarea,
[data-testid="stSidebar"] .stTextArea label,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] {{
  font-size: 0.8rem !important;
  letter-spacing: 0.04em !important;
  line-height: 1.7 !important;
}}
[data-testid="stSidebar"] textarea {{
  font-size: 0.8rem !important;
  line-height: 1.65 !important;
}}

/* ── 頁面標題 ── */
h1 {{
  font-size: 1.8rem !important;
  font-weight: 600 !important;
  color: #1A1D2E !important;
  letter-spacing: 0.12em !important;
  line-height: 1.4 !important;
  margin-bottom: 0 !important;
  background: linear-gradient(135deg, #3A6EC8 0%, #7B5EA7 55%, #3AA8C4 100%);
  -webkit-background-clip: text !important;
  -webkit-text-fill-color: transparent !important;
  background-clip: text !important;
}}
[data-testid="stCaptionContainer"] p {{
  font-size: 0.95rem !important;
  color: #8899BB !important;
  letter-spacing: 0.06em !important;
  line-height: 1.9 !important;
  font-weight: 300 !important;
}}

/* ══════════════════════════════════════
   主要按鈕 — 玻璃漸層藍
   ══════════════════════════════════════ */
div.stButton > button[kind="primary"],
div.stDownloadButton > button[kind="primary"] {{
  background: linear-gradient(135deg, #5B8FD4 0%, #7B6EC8 100%) !important;
  border: none !important;
  color: #ffffff !important;
  border-radius: 12px !important;
  font-weight: 500 !important;
  font-size: 0.88rem !important;
  letter-spacing: 0.1em !important;
  line-height: 1.5 !important;
  padding: 0.5rem 1.2rem !important;
  box-shadow: 0 4px 15px rgba(91,143,212,0.30), 0 1px 4px rgba(0,0,0,0.08) !important;
  transition: all 0.2s !important;
  backdrop-filter: blur(8px) !important;
}}
div.stButton > button[kind="primary"]:hover,
div.stDownloadButton > button[kind="primary"]:hover {{
  background: linear-gradient(135deg, #4A7EC3 0%, #6A5DB7 100%) !important;
  box-shadow: 0 6px 20px rgba(91,143,212,0.40), 0 2px 6px rgba(0,0,0,0.10) !important;
  transform: translateY(-1px) !important;
}}

/* ── 次要按鈕 — 霧玻璃 ── */
div.stButton > button:not([kind="primary"]) {{
  border-radius: 10px !important;
  font-size: 0.85rem !important;
  letter-spacing: 0.05em !important;
  line-height: 1.5 !important;
  border: 1px solid rgba(160,185,240,0.45) !important;
  color: #4A6090 !important;
  background: rgba(255,255,255,0.68) !important;
  backdrop-filter: blur(10px) saturate(1.3) !important;
  -webkit-backdrop-filter: blur(10px) saturate(1.3) !important;
  box-shadow: 0 2px 8px rgba(91,143,212,0.10), inset 0 1px 0 rgba(255,255,255,0.8) !important;
  transition: all 0.18s !important;
}}
div.stButton > button:not([kind="primary"]):hover {{
  background: rgba(235,242,255,0.85) !important;
  border-color: rgba(91,143,212,0.5) !important;
  box-shadow: 0 4px 12px rgba(91,143,212,0.18) !important;
  transform: translateY(-1px) !important;
}}

/* ══════════════════════════════════════
   區塊標題
   ══════════════════════════════════════ */
.section-title {{
  font-size: 1.05rem;
  font-weight: 600;
  color: #2A3A5C;
  margin-bottom: 0.15rem;
  letter-spacing: 0.15em;
  line-height: 1.6;
  border-left: 3px solid transparent;
  border-image: linear-gradient(to bottom, #5B8FD4, #9B6EC8) 1;
  padding-left: 10px;
}}
.section-sub {{
  color: #8899BB;
  font-size: 0.88rem;
  font-weight: 300;
  margin-bottom: 0.85rem;
  line-height: 1.9;
  letter-spacing: 0.06em;
  padding-left: 13px;
}}

/* ══════════════════════════════════════
   Card 容器 — 液態玻璃主視覺
   ══════════════════════════════════════ */
[data-testid="stVerticalBlock"] > [data-testid="element-container"] > div[style*="border"] {{
  border-radius: 18px !important;
  border: 1px solid rgba(160,190,255,0.40) !important;
  border-top: 1px solid rgba(200,215,255,0.65) !important;
  background: rgba(255,255,255,0.65) !important;
  backdrop-filter: blur(20px) saturate(1.6) !important;
  -webkit-backdrop-filter: blur(20px) saturate(1.6) !important;
  box-shadow:
    0 8px 32px rgba(91,143,212,0.10),
    0 2px 8px rgba(0,0,0,0.05),
    inset 0 1px 0 rgba(255,255,255,0.85),
    inset 0 -1px 0 rgba(160,190,255,0.15) !important;
}}

/* ══════════════════════════════════════
   Timeline wrapper — 玻璃面板
   ══════════════════════════════════════ */
.timeline-wrap {{
  overflow-x: auto;
  border: 1px solid rgba(160,190,255,0.35);
  border-radius: 14px;
  background: rgba(248,251,255,0.80);
  backdrop-filter: blur(16px) !important;
  -webkit-backdrop-filter: blur(16px) !important;
  box-shadow: 0 4px 20px rgba(91,143,212,0.09), inset 0 1px 0 rgba(255,255,255,0.9);
  margin-top: 8px;
}}

/* ── Timeline table base ── */
.timeline-table {{
  border-collapse: collapse;
  width: max-content;
  min-width: 100%;
  font-size: 12px;
  font-weight: 300;
}}
.timeline-table th,
.timeline-table td {{
  border: 1px solid rgba(180,200,240,0.35);
  text-align: center;
  padding: 0;
  height: 33px;
}}

/* ── 月份 header：淡彩玻璃 ── */
.timeline-table .month-row th {{
  height: 25px;
  background: rgba(220,230,255,0.55);
  font-weight: 600;
  font-size: 10.5px;
  color: #5566AA;
  letter-spacing: 2px;
  text-transform: uppercase;
  backdrop-filter: blur(8px);
}}

/* ── 日期 & 星期 header ── */
.timeline-table .date-head {{
  width: 32px; min-width: 32px; max-width: 32px;
  font-size: 10.5px;
  line-height: 1.2;
  color: #6677AA;
  background: rgba(235,242,255,0.60);
  font-weight: 300;
}}
.timeline-table .weekend-head {{
  background: rgba(215,225,250,0.55) !important;
  color: #99AACC !important;
}}
.timeline-table .weekend-cell {{
  background: rgba(235,240,255,0.35);
}}
.timeline-table .empty-cell {{
  background: rgba(248,251,255,0.50);
}}

/* ── 固定左欄 ── */
.timeline-table .task-col {{
  min-width: 186px; max-width: 186px; width: 186px;
  text-align: left; padding: 0 14px;
  font-weight: 400; font-size: 12px;
  background: rgba(248,251,255,0.88);
  position: sticky; left: 0; z-index: 3;
  border-right: 1px solid rgba(160,185,240,0.35);
  color: #1A2540;
  letter-spacing: 0.4px;
  backdrop-filter: blur(12px);
}}
.timeline-table .owner-col {{
  min-width: 90px; max-width: 90px; width: 90px;
  background: rgba(248,251,255,0.88);
  position: sticky; left: 186px; z-index: 3;
  font-size: 11.5px;
  color: #7788AA;
  border-right: 1px solid rgba(160,185,240,0.35);
  font-weight: 300;
  backdrop-filter: blur(12px);
}}

/* Sticky header cells */
.timeline-table .month-row .task-col,
.timeline-table .month-row .owner-col,
tr:nth-child(2) .task-col,
tr:nth-child(2) .owner-col,
tr:nth-child(3) .task-col,
tr:nth-child(3) .owner-col {{
  background: rgba(220,230,255,0.75);
}}

/* ── BREAK column ── */
.timeline-table .break-cell {{
  width: 18px; min-width: 18px; max-width: 18px;
  background: linear-gradient(to bottom, rgba(200,215,255,0.5), rgba(220,200,255,0.5));
  color: #99AACC;
  font-weight: 400;
  font-size: 11px;
  writing-mode: vertical-rl;
  text-orientation: mixed;
  letter-spacing: 4px;
  vertical-align: middle;
  border-left: 1px solid rgba(160,185,240,0.3);
  border-right: 1px solid rgba(160,185,240,0.3);
}}

/* ── Bar cells — 彩色半透明 ── */
.timeline-table .bar-ad2    {{
  background: linear-gradient(135deg, rgba(91,164,212,0.85) 0%, rgba(100,140,220,0.85) 100%);
  border-color: rgba(91,143,212,0.25);
}}
.timeline-table .bar-client {{
  background: linear-gradient(135deg, rgba(232,149,106,0.85) 0%, rgba(240,130,100,0.85) 100%);
  border-color: rgba(220,100,80,0.20);
}}
.timeline-table .bar-launch {{
  background: linear-gradient(135deg, rgba(224,92,122,0.90) 0%, rgba(200,70,140,0.85) 100%);
  border-color: rgba(200,70,120,0.20);
}}
.timeline-table .bar-prep   {{
  background: linear-gradient(135deg, rgba(110,196,160,0.85) 0%, rgba(80,185,170,0.85) 100%);
  border-color: rgba(80,185,160,0.20);
}}
.timeline-table tr.separator-row td {{
  border-bottom: 3px solid rgba(91,143,212,0.55) !important;
}}

/* ── Legend ── */
.legend {{
  display: flex; gap: 20px; flex-wrap: wrap; margin-bottom: 12px;
  font-size: 11px; color: {UI_MUTED};
  align-items: center;
  padding: 4px 0;
  letter-spacing: 0.8px;
}}
.legend-item {{ display: inline-flex; align-items: center; gap: 6px; }}
.legend-dot {{
  width: 10px; height: 10px; border-radius: 3px; display: inline-block;
  opacity: 0.9;
  box-shadow: 0 1px 4px rgba(0,0,0,0.15);
}}

/* ── 任務清單標題列 ── */
.task-head-label {{
  font-size: 11px;
  font-weight: 400;
  color: #99AACC;
  letter-spacing: 0.12em;
  line-height: 1.8;
  padding: 2px 0;
  display: block;
}}
.task-head-center {{
  text-align: center;
  display: flex;
  justify-content: center;
}}

/* ── 斑馬條紋 ── */
.task-row-plain     {{ background: transparent; }}
.task-row-plain-alt {{ background: rgba(235,242,255,0.40); border-radius: 4px; }}

/* ── 操作按鈕 ── */
.op-btn button {{
  font-size: 12px !important;
  padding: 0 !important;
  height: 2.1rem !important;
  min-height: 2.1rem !important;
  border-radius: 8px !important;
}}

/* ── 任務行分隔線 ── */
.flow-row-sep {{
  height: 1px;
  background: linear-gradient(to right, rgba(160,185,240,0.4) 0%, rgba(200,215,255,0.2) 60%, transparent 100%);
  margin: 0.2rem 0 0.35rem 0;
}}

/* ══════════════════════════════════════
   Input 共用 — 玻璃質感
   ══════════════════════════════════════ */
[data-testid="stCheckbox"] {{
  margin-top: 0 !important;
  margin-bottom: 0 !important;
}}
[data-testid="stCheckbox"] label {{
  min-height: auto !important;
}}
[data-testid="stTextInput"],
[data-testid="stNumberInput"],
[data-testid="stSelectbox"],
[data-testid="stCheckbox"],
div.stButton {{
  margin-top: 0 !important;
  margin-bottom: 0 !important;
}}
[data-testid="stTextInput"] > div,
[data-testid="stNumberInput"] > div,
[data-testid="stSelectbox"] > div {{
  margin-top: 0 !important;
  margin-bottom: 0 !important;
}}
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {{
  border-radius: 10px !important;
  border: 1px solid rgba(160,185,240,0.45) !important;
  background: rgba(255,255,255,0.72) !important;
  backdrop-filter: blur(12px) !important;
  -webkit-backdrop-filter: blur(12px) !important;
  font-size: 0.92rem !important;
  font-weight: 300 !important;
  height: 2.6rem !important;
  color: #1A2540 !important;
  letter-spacing: 0.05em !important;
  line-height: 1.6 !important;
  box-shadow: 0 2px 8px rgba(91,143,212,0.08), inset 0 1px 0 rgba(255,255,255,0.9) !important;
}}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {{
  border-color: rgba(91,143,212,0.65) !important;
  box-shadow: 0 0 0 3px rgba(91,143,212,0.15), 0 2px 8px rgba(91,143,212,0.12) !important;
  background: rgba(255,255,255,0.92) !important;
}}
[data-testid="stSelectbox"] > div > div,
[data-baseweb="select"] > div {{
  min-height: 2.6rem !important;
  border-radius: 10px !important;
  border: 1px solid rgba(160,185,240,0.45) !important;
  background: rgba(255,255,255,0.72) !important;
  backdrop-filter: blur(12px) !important;
  box-shadow: 0 2px 8px rgba(91,143,212,0.08), inset 0 1px 0 rgba(255,255,255,0.9) !important;
}}
[data-testid="stSelectbox"] [data-baseweb="select"] {{
  margin: 0 !important;
}}
[data-testid="stDateInput"] input {{
  border-radius: 10px !important;
  border: 1px solid rgba(160,185,240,0.45) !important;
  background: rgba(255,255,255,0.72) !important;
}}
[data-testid="column"] > div[data-testid="stVerticalBlock"] {{
  gap: 0.2rem !important;
}}
[data-testid="element-container"] {{
  margin-bottom: 0.2rem !important;
}}

/* ── 間距 ── */
.small-gap {{ height: 0.3rem; }}
.large-gap {{ height: 1.6rem; }}

/* ── 訊息提示 ── */
[data-testid="stAlert"] {{
  border-radius: 12px !important;
  border-left-width: 3px !important;
  background: rgba(255,255,255,0.72) !important;
  backdrop-filter: blur(10px) !important;
  font-size: 0.85rem !important;
  font-weight: 300 !important;
  letter-spacing: 0.3px !important;
  box-shadow: 0 2px 10px rgba(91,143,212,0.08) !important;
}}

/* ── Tabs 玻璃風格 ── */
[data-testid="stTabs"] [role="tab"] {{
  border-radius: 8px 8px 0 0 !important;
  background: rgba(255,255,255,0.45) !important;
  border: 1px solid rgba(160,185,240,0.30) !important;
  border-bottom: none !important;
  color: #6688AA !important;
  font-size: 0.88rem !important;
  font-weight: 400 !important;
  letter-spacing: 0.08em !important;
}}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {{
  background: rgba(255,255,255,0.80) !important;
  color: {UI_PRIMARY} !important;
  font-weight: 600 !important;
  border-color: rgba(91,143,212,0.35) !important;
  box-shadow: 0 -2px 8px rgba(91,143,212,0.10) !important;
}}

/* ── Checkbox 色彩 ── */
[data-testid="stCheckbox"] input:checked + div {{
  background: {UI_PRIMARY} !important;
  border-color: {UI_PRIMARY} !important;
}}
</style>
""", unsafe_allow_html=True)

# =========================
# state
# =========================
def init_state():
    if "project_name" not in st.session_state:
        st.session_state.project_name = ""
    if "mode_display" not in st.session_state:
        st.session_state.mode_display = MODE_OPTIONS[0]
    if "start_date_value" not in st.session_state:
        st.session_state.start_date_value = date.today()
    if "launch_date_value" not in st.session_state:
        st.session_state.launch_date_value = date.today()
    if "collapse_threshold" not in st.session_state:
        st.session_state.collapse_threshold = 2
    if "tasks" not in st.session_state:
        st.session_state.tasks = [row.copy() for row in DEFAULT_TASKS]
    if "holidays_text" not in st.session_state:
        st.session_state.holidays_text = "\n".join(f"{k},{v}" for k, v in DEFAULT_HOLIDAYS.items())
    if "schedule_df" not in st.session_state:
        st.session_state.schedule_df = None
    if "excel_bytes" not in st.session_state:
        st.session_state.excel_bytes = None
    if "display_columns" not in st.session_state:
        st.session_state.display_columns = None
    if "holidays_dt" not in st.session_state:
        st.session_state.holidays_dt = None
    if "warning_msg" not in st.session_state:
        st.session_state.warning_msg = ""
    if "last_generated_name" not in st.session_state:
        st.session_state.last_generated_name = "未命名專案"
    if "status_msg" not in st.session_state:
        st.session_state.status_msg = ""
    if "validation_error_msg" not in st.session_state:
        st.session_state.validation_error_msg = ""
    if "batch_tasks_text" not in st.session_state:
        st.session_state.batch_tasks_text = DEFAULT_BATCH_TASKS_TEXT
    if "batch_template_display" not in st.session_state:
        st.session_state.batch_template_display = BATCH_TEMPLATE_OPTIONS[0]
    if "batch_msg" not in st.session_state:
        st.session_state.batch_msg = ""
    if "import_msg" not in st.session_state:
        st.session_state.import_msg = ""
    if "half_day_label" not in st.session_state:
        st.session_state.half_day_label = "1300"

init_state()

def ensure_task_ids():
    tasks = st.session_state.tasks
    existing = set()
    for i, row in enumerate(tasks):
        rid = row.get("id")
        if not rid or rid in existing:
            row["id"] = f"task_{i+1}_{uuid.uuid4().hex[:6]}"
        existing.add(row["id"])

ensure_task_ids()

# =========================
# helpers
# =========================
def parse_holidays(text: str) -> dict:
    holidays = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "," not in line:
            raise ValueError(f"假日格式錯誤：{line}，請使用 YYYY-MM-DD,名稱")
        d, name = line.split(",", 1)
        pd.to_datetime(d.strip())
        holidays[d.strip()] = name.strip()
    return holidays

def get_active_tasks():
    tasks = []
    visible_rows = [row for row in st.session_state.tasks if row.get("顯示", True) and str(row.get("任務名稱", "")).strip()]
    if not visible_rows:
        raise ValueError("目前沒有可排程的任務，請至少保留一筆顯示中的任務。")

    launch_count = sum(bool(r.get("上線日", False)) for r in visible_rows)
    if launch_count > 1:
        raise ValueError("「上線日」只能勾選一筆。")
    if launch_count == 0:
        visible_rows[-1]["上線日"] = True

    for row in visible_rows:
        try:
            days = float(row.get("工作天數", 0) or 0)
        except (TypeError, ValueError):
            raise ValueError(f"任務「{row.get('任務名稱','未命名')}」的工作天數格式錯誤。")
        if days < 0.5:
            raise ValueError(f"任務「{row.get('任務名稱','未命名')}」的工作天數需至少 0.5 天。")
        if abs(days * 2 - round(days * 2)) > 1e-9:
            raise ValueError(f"任務「{row.get('任務名稱','未命名')}」的工作天數需以 0.5 天為單位。")
        half_label = str(row.get("半天標註", DEFAULT_HALF_DAY_LABEL) or DEFAULT_HALF_DAY_LABEL).strip() or DEFAULT_HALF_DAY_LABEL
        tasks.append({
            "task": str(row.get("任務名稱", "")).strip(),
            "owner": str(row.get("Action By", "Ad2")).strip() or "Ad2",
            "days": days,
            "half_day_label": half_label,
            "is_launch": bool(row.get("上線日", False)),
            "thick_bottom": bool(row.get("粗下框線", False)),
        })
    return tasks

def build_scheduler(tasks_config, holidays_config, calculation_mode, start_date_obj, launch_date_obj):
    holidays_dt = [pd.to_datetime(h).date() for h in holidays_config.keys()]

    def is_workday(d):
        return (d.weekday() < 5) and (d not in holidays_dt)

    def ceil_day_units(days):
        return int(round(float(days) * 2))

    def format_days(days):
        days = float(days)
        return int(days) if days.is_integer() else days

    def add_workdays(start_date, days):
        current = start_date
        check_days = max(int(days) - 1, 0)
        while check_days > 0:
            current += timedelta(days=1)
            if is_workday(current):
                check_days -= 1
        return current

    def get_previous_workday(d):
        d -= timedelta(days=1)
        while not is_workday(d):
            d -= timedelta(days=1)
        return d

    def get_next_workday(d):
        d += timedelta(days=1)
        while not is_workday(d):
            d += timedelta(days=1)
        return d

    def ensure_workday_forward(d):
        while not is_workday(d):
            d += timedelta(days=1)
        return d

    def ensure_workday_backward(d):
        while not is_workday(d):
            d -= timedelta(days=1)
        return d

    def advance_slot(slot, half_units):
        """slot=(date, 0/1)，0=上午、1=下午；half_units 可正可負。"""
        current_date, half = slot
        step = 1 if half_units >= 0 else -1
        for _ in range(abs(int(half_units))):
            if step > 0:
                if half == 0:
                    half = 1
                else:
                    current_date = get_next_workday(current_date)
                    half = 0
            else:
                if half == 1:
                    half = 0
                else:
                    current_date = get_previous_workday(current_date)
                    half = 1
        return current_date, half

    def schedule_row(t, start_slot):
        units = ceil_day_units(t["days"])
        end_slot = advance_slot(start_slot, units - 1)
        return {
            "Task": t["task"],
            "Owner": t["owner"],
            "Start Date": start_slot[0],
            "End Date": end_slot[0],
            "Start Half": start_slot[1],
            "End Half": end_slot[1],
            "Duration Days": float(t["days"]),
            "Half Day Label": t.get("half_day_label", DEFAULT_HALF_DAY_LABEL),
            "Half Units": units,
            "Type": "Launch" if t["is_launch"] else "Normal",
            "Thick Bottom": bool(t.get("thick_bottom", False)),
        }, advance_slot(start_slot, units)

    schedule = []
    warning_msg = ""
    launch_task_config = next((t for t in tasks_config if t["is_launch"]), tasks_config[-1])

    if calculation_mode == "backward":
        if not launch_date_obj:
            raise ValueError("「上線日回推」需要填寫上線日期。")
        current_end_slot = (ensure_workday_backward(launch_date_obj), 1)
        temp_schedule = []
        for t in tasks_config[::-1]:
            units = ceil_day_units(t["days"])
            start_slot = advance_slot(current_end_slot, -(units - 1))
            row = {
                "Task": t["task"],
                "Owner": t["owner"],
                "Start Date": start_slot[0],
                "End Date": current_end_slot[0],
                "Start Half": start_slot[1],
                "End Half": current_end_slot[1],
                "Duration Days": float(t["days"]),
                "Half Units": units,
                "Type": "Launch" if t["is_launch"] else "Normal",
                "Thick Bottom": bool(t.get("thick_bottom", False)),
            }
            temp_schedule.append(row)
            current_end_slot = advance_slot(start_slot, -1)
        schedule = temp_schedule[::-1]

    elif calculation_mode == "forward":
        curr_start_date = ensure_workday_forward(start_date_obj or date.today())
        curr_slot = (curr_start_date, 0)
        for t in tasks_config:
            if t["is_launch"] and launch_date_obj:
                curr_slot = (ensure_workday_forward(launch_date_obj), 0)
            row, curr_slot = schedule_row(t, curr_slot)
            schedule.append(row)

    else:
        if not start_date_obj or not launch_date_obj:
            raise ValueError("「同時指定開始與上線日期」需要同時填寫開始日期與上線日期。")
        curr_slot = (ensure_workday_forward(start_date_obj), 0)
        normal_tasks = [t for t in tasks_config if not t["is_launch"]]
        for t in normal_tasks:
            row, curr_slot = schedule_row(t, curr_slot)
            schedule.append(row)

        last_task_end = schedule[-1]["End Date"] if schedule else start_date_obj
        real_prep_start = get_next_workday(last_task_end)
        real_prep_end = launch_date_obj - timedelta(days=1)

        if last_task_end >= launch_date_obj:
            overrun_days = (last_task_end - launch_date_obj).days
            warning_msg = f"⚠️【時程衝突警告】工作將進行到 {last_task_end}，比上線日晚了 {overrun_days} 天。"

        if real_prep_end >= real_prep_start:
            schedule.append({
                "Task": "預備上線", "Owner": "Ad2",
                "Start Date": real_prep_start, "End Date": real_prep_end,
                "Start Half": 0, "End Half": 1,
                "Duration Days": None, "Half Units": None,
                "Type": "Prep",
                "Thick Bottom": False
            })

        launch_row, _ = schedule_row(launch_task_config, (ensure_workday_forward(launch_date_obj), 0))
        schedule.append(launch_row)

    return pd.DataFrame(schedule), warning_msg, holidays_dt

def prepare_display_columns(df_schedule, holidays_dt, launch_date_obj, collapse_threshold):
    min_date = df_schedule["Start Date"].min()
    max_date = df_schedule["End Date"].max()
    if launch_date_obj and launch_date_obj > max_date:
        max_date = launch_date_obj
    full_dates = list(pd.date_range(min_date, max_date, freq="D"))

    display_columns = []
    prep_task_row = df_schedule[df_schedule["Type"] == "Prep"]
    prep_task = None
    if not prep_task_row.empty:
        r = prep_task_row.iloc[0]
        prep_task = {"Start Date": r["Start Date"], "End Date": r["End Date"]}

    if prep_task and (prep_task["End Date"] - prep_task["Start Date"]).days + 1 >= collapse_threshold:
        keep_start = prep_task["Start Date"]
        resume_date = prep_task["End Date"] + timedelta(days=1)
        break_added = False
        for d in full_dates:
            d_date = d.date()
            if d_date >= resume_date or d_date <= keep_start:
                display_columns.append(d)
            else:
                if not break_added:
                    display_columns.append("BREAK")
                    break_added = True
    else:
        display_columns = full_dates

    return display_columns

def compute_month_segments(display_columns, col_start):
    segments = []
    segment_start_col = None
    segment_month = None
    segment_year = None

    for i, item in enumerate(display_columns):
        if item == "BREAK":
            if segment_start_col is not None:
                segments.append((segment_start_col, col_start + i - 1, segment_year, segment_month))
                segment_start_col = None
                segment_month = None
                segment_year = None
            continue

        d = item
        excel_col = col_start + i
        if segment_start_col is None:
            segment_start_col = excel_col
            segment_month = d.month
            segment_year = d.year
        elif d.month != segment_month or d.year != segment_year:
            segments.append((segment_start_col, excel_col - 1, segment_year, segment_month))
            segment_start_col = excel_col
            segment_month = d.month
            segment_year = d.year

    if segment_start_col is not None:
        last_real_col = None
        for i in range(len(display_columns) - 1, -1, -1):
            if display_columns[i] != "BREAK":
                last_real_col = col_start + i
                break
        if last_real_col is not None:
            segments.append((segment_start_col, last_real_col, segment_year, segment_month))

    return segments

def build_excel_bytes(df_schedule, holidays_config, holidays_dt, launch_date_obj, collapse_threshold):
    output = io.BytesIO()
    display_columns = prepare_display_columns(df_schedule, holidays_dt, launch_date_obj, collapse_threshold)

    def is_workday(d):
        return (d.weekday() < 5) and (d not in holidays_dt)

    holiday_blocks_info = []
    current_block_start = -1
    current_block_dates = []

    for i, col_item in enumerate(display_columns):
        is_holiday_day = False
        if col_item != "BREAK":
            d_date = col_item.date()
            if not is_workday(d_date):
                is_holiday_day = False if (launch_date_obj and d_date == launch_date_obj) else True

        if is_holiday_day:
            if current_block_start == -1:
                current_block_start = i
            current_block_dates.append(col_item.date())
        else:
            if current_block_start != -1:
                found_name = next((holidays_config[d.strftime("%Y-%m-%d")] for d in current_block_dates if d.strftime("%Y-%m-%d") in holidays_config), None)
                if found_name:
                    holiday_blocks_info.append({"start_col": current_block_start, "end_col": i - 1, "name": "\n".join(list(found_name))})
                elif len(current_block_dates) > 4:
                    holiday_blocks_info.append({"start_col": current_block_start, "end_col": i - 1, "name": "長\n假"})
            current_block_start, current_block_dates = -1, []
    if current_block_start != -1:
        found_name = next((holidays_config[d.strftime("%Y-%m-%d")] for d in current_block_dates if d.strftime("%Y-%m-%d") in holidays_config), None)
        if found_name:
            holiday_blocks_info.append({"start_col": current_block_start, "end_col": len(display_columns) - 1, "name": "\n".join(list(found_name))})

    writer = pd.ExcelWriter(output, engine="xlsxwriter")
    workbook = writer.book
    worksheet = workbook.add_worksheet("時程表")
    font = "Microsoft JhengHei"
    FONT_SIZE = 11
    border_fmt = {"border": 1, "border_color": "#000000"}

    def F(**kwargs):
        return workbook.add_format({"font_name": font, "font_size": FONT_SIZE, **kwargs})

    fmt_center = F(align="center", valign="vcenter", **border_fmt)
    fmt_left = F(align="left", valign="vcenter", **border_fmt)
    fmt_weekend = F(bg_color=EXCEL_COLOR_WEEKEND, align="center", valign="vcenter", **border_fmt)
    fmt_date_num = F(align="center", valign="vcenter", **border_fmt)
    fmt_holiday_merged = F(align="center", valign="vcenter", text_wrap=True, bg_color=EXCEL_COLOR_WEEKEND, border=1, font_color=EXCEL_COLOR_HOLIDAY_TEXT, bold=True)
    fmt_header_main = F(bold=True, align="center", valign="vcenter", bg_color="#FFFFFF", **border_fmt)
    fmt_bar_client = F(bg_color=EXCEL_COLOR_CLIENT_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_bar_ad2 = F(bg_color=EXCEL_COLOR_AD2_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_bar_launch = F(bg_color=EXCEL_COLOR_LAUNCH_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_bar_prep = F(bg_color=EXCEL_COLOR_PREP_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_legend_client = F(bg_color=EXCEL_COLOR_CLIENT_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_legend_ad2 = F(bg_color=EXCEL_COLOR_AD2_BAR, align="center", valign="vcenter", **border_fmt)
    fmt_break_merge = F(align="center", valign="vcenter", **border_fmt)

    worksheet.write(0, 2, "客戶", fmt_legend_client)
    worksheet.write(0, 3, "Ad2", fmt_legend_ad2)
    # 上方色票只保留單獨角色，避免匯出時多出 Ad2＋客戶 標籤。
    worksheet.merge_range(1, 0, 3, 0, "製作時程", fmt_header_main)
    worksheet.merge_range(1, 1, 3, 1, "Action by", fmt_header_main)
    worksheet.set_column(0, 0, 20)
    worksheet.set_column(1, 1, 12)

    col_start, row_start = 2, 4
    # 固定 Excel 畫面：上方月份／日期標題列 + 左側 A/B 欄（製作時程、Action by）
    worksheet.freeze_panes(row_start, col_start)
    break_cols_excel = []

    month_segments = compute_month_segments(display_columns, col_start)
    for start_col, end_col, year, month in month_segments:
        month_fmt = F(bold=True, align="center", valign="vcenter", bg_color=get_month_color(month), **border_fmt)
        month_label = date(year, month, 1).strftime("%b").upper()
        if start_col == end_col:
            worksheet.write(1, start_col, month_label, month_fmt)
        else:
            worksheet.merge_range(1, start_col, 1, end_col, month_label, month_fmt)

    for i, item in enumerate(display_columns):
        col = col_start + i
        if item == "BREAK":
            worksheet.set_column(col, col, 4)
            break_cols_excel.append(col)
            continue
        d = item
        is_h = not is_workday(d.date())
        weekday_map = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五", 5: "六", 6: "日"}
        worksheet.write(2, col, d.day, fmt_weekend if is_h else fmt_date_num)
        worksheet.write(3, col, weekday_map[d.weekday()], fmt_weekend if is_h else fmt_date_num)
        worksheet.set_column(col, col, 4.5)

    last_task_row = row_start + len(df_schedule) - 1
    for c in break_cols_excel:
        worksheet.merge_range(1, c, last_task_row, c, "～", fmt_break_merge)

    for idx, item in df_schedule.iterrows():
        row = row_start + idx
        worksheet.write(row, 0, item["Task"], fmt_left)
        worksheet.write(row, 1, item["Owner"], fmt_left)
        if item["Type"] == "Launch":
            bar_fmt = fmt_bar_launch
        elif item["Type"] == "Prep":
            bar_fmt = fmt_bar_prep
        elif item["Owner"] == "Ad2＋客戶":
            bar_fmt = fmt_bar_prep
        elif "客戶" in item["Owner"]:
            bar_fmt = fmt_bar_client
        else:
            bar_fmt = fmt_bar_ad2

        for i, col_item in enumerate(display_columns):
            if col_item == "BREAK":
                continue
            col = col_start + i
            d_date = col_item.date()
            if item["Start Date"] <= d_date <= item["End Date"]:
                if item["Type"] in ["Launch", "Prep"] or is_workday(d_date):
                    worksheet.write(row, col, str(item.get("Half Day Label", DEFAULT_HALF_DAY_LABEL) or DEFAULT_HALF_DAY_LABEL) if (abs(float(item.get("Duration Days") or 0) % 1 - 0.5) < 1e-9 and d_date == item["End Date"]) else "", bar_fmt)
                else:
                    worksheet.write(row, col, "", fmt_weekend)
            else:
                worksheet.write(row, col, "", fmt_weekend if not is_workday(d_date) else fmt_center)

    for block in holiday_blocks_info:
        c1, c2 = col_start + block["start_col"], col_start + block["end_col"]
        if c1 == c2:
            worksheet.merge_range(row_start, c1, last_task_row, c1, block["name"], fmt_holiday_merged)
        else:
            worksheet.merge_range(row_start, c1, last_task_row, c2, block["name"], fmt_holiday_merged)

    writer.close()

    # 批次輸入若以「--」標記階段分隔，會在上一個任務列套用粗下框線。
    thick_rows = []
    if "Thick Bottom" in df_schedule.columns:
        for idx, item in df_schedule.iterrows():
            if bool(item.get("Thick Bottom", False)):
                thick_rows.append(row_start + int(idx) + 1)  # openpyxl row is 1-based

    if thick_rows:
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["時程表"]
        thick_side = Side(style="thick", color="000000")
        last_col = ws.max_column
        for excel_row in thick_rows:
            for col in range(1, last_col + 1):
                cell = ws.cell(row=excel_row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                border = copy(cell.border)
                border.bottom = thick_side
                cell.border = border
        patched_output = io.BytesIO()
        wb.save(patched_output)
        patched_output.seek(0)
        return patched_output.getvalue(), display_columns

    output.seek(0)
    return output.getvalue(), display_columns

def render_stable_preview(df_schedule, display_columns, holidays_dt):
    def is_workday(d):
        return (d.weekday() < 5) and (d not in holidays_dt)
    weekday_map = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五", 5: "六", 6: "日"}

    # 3 header rows (month / date / weekday) + all task rows
    total_rowspan = 3 + len(df_schedule)

    month_cells = []
    i = 0
    while i < len(display_columns):
        item = display_columns[i]
        if item == "BREAK":
            # Merge entire column (including month/date/weekday headers + all task rows)
            month_cells.append(f'<th class="break-cell" rowspan="{total_rowspan}">～</th>')
            i += 1
            continue
        month = item.strftime("%m月")
        span = 1
        j = i + 1
        while j < len(display_columns):
            nxt = display_columns[j]
            if nxt == "BREAK" or nxt.strftime("%m") != item.strftime("%m"):
                break
            span += 1
            j += 1
        month_cells.append(f'<th colspan="{span}" style="background:{get_month_color(item.month)};">{month}</th>')
        i = j

    date_cells = []
    weekday_cells = []
    for item in display_columns:
        if item == "BREAK":
            continue  # already covered by rowspan above
        d = item.date()
        cls = "date-head weekend-head" if not is_workday(d) else "date-head"
        date_cells.append(f'<th class="{cls}">{d.day}</th>')
        weekday_cells.append(f'<th class="{cls}">{weekday_map[item.weekday()]}</th>')

    rows = []
    rows.append(
        "<tr class='month-row'>"
        "<th class='task-col'>任務名稱</th>"
        "<th class='owner-col'>Action By</th>"
        + "".join(month_cells) + "</tr>"
    )
    rows.append("<tr><th class='task-col'></th><th class='owner-col'></th>" + "".join(date_cells) + "</tr>")
    rows.append("<tr><th class='task-col'></th><th class='owner-col'></th>" + "".join(weekday_cells) + "</tr>")

    for _, row in df_schedule.iterrows():
        cells = [
            f'<td class="task-col">{row["Task"]}</td>',
            f'<td class="owner-col">{row["Owner"]}</td>',
        ]
        for item in display_columns:
            if item == "BREAK":
                continue  # already merged by rowspan — skip this cell
            d = item.date()
            base_cls = "weekend-cell" if not is_workday(d) else "empty-cell"
            if row["Start Date"] <= d <= row["End Date"]:
                if row["Type"] in ["Launch", "Prep"] or is_workday(d):
                    if row["Type"] == "Launch":
                        cls = "bar-launch"
                    elif row["Type"] == "Prep":
                        cls = "bar-prep"
                    elif row["Owner"] == "Ad2＋客戶":
                        cls = "bar-prep"
                    elif "客戶" in row["Owner"]:
                        cls = "bar-client"
                    else:
                        cls = "bar-ad2"
                    cells.append(f'<td class="{cls}">{str(row.get("Half Day Label", DEFAULT_HALF_DAY_LABEL) or DEFAULT_HALF_DAY_LABEL) if (abs(float(row.get("Duration Days") or 0) % 1 - 0.5) < 1e-9 and d == row["End Date"]) else ""}</td>')
                else:
                    cells.append(f'<td class="{base_cls}"></td>')
            else:
                cells.append(f'<td class="{base_cls}"></td>')
        row_class = " class='separator-row'" if bool(row.get("Thick Bottom", False)) else ""
        rows.append(f"<tr{row_class}>" + "".join(cells) + "</tr>")

    return f"""
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:{UI_AD2};"></span>Ad2</span>
      <span class="legend-item"><span class="legend-dot" style="background:{UI_CLIENT};"></span>客戶</span>
      <span class="legend-item"><span class="legend-dot" style="background:{UI_LAUNCH};"></span>上線</span>
    </div>
    <div class="timeline-wrap">
      <table class="timeline-table">
        {''.join(rows)}
      </table>
    </div>
    """

def sync_task_field(task_id: str, field: str, widget_key: str):
    for row in st.session_state.tasks:
        if row.get("id") == task_id:
            row[field] = st.session_state.get(widget_key)
            break

def sync_launch_field(task_id: str, widget_key: str):
    """讓上線日像單選題一樣運作，避免誤選兩筆造成產出錯誤。"""
    selected = bool(st.session_state.get(widget_key))

    for row in st.session_state.tasks:
        rid = row.get("id")
        is_target = rid == task_id
        row["上線日"] = selected if is_target else False

        launch_key = f"launch_{rid}"
        if launch_key in st.session_state:
            st.session_state[launch_key] = selected if is_target else False

    st.session_state.validation_error_msg = ""

def add_task():
    st.session_state.tasks.append({
        "id": f"task_new_{uuid.uuid4().hex[:6]}",
        "顯示": True,
        "任務名稱": "",
        "Action By": "Ad2",
        "工作天數": 1.0,
        "半天標註": DEFAULT_HALF_DAY_LABEL,
        "上線日": False,
        "粗下框線": False,
    })

def move_task_up(idx: int):
    if 0 < idx < len(st.session_state.tasks):
        st.session_state.tasks[idx - 1], st.session_state.tasks[idx] = st.session_state.tasks[idx], st.session_state.tasks[idx - 1]

def move_task_down(idx: int):
    if 0 <= idx < len(st.session_state.tasks) - 1:
        st.session_state.tasks[idx + 1], st.session_state.tasks[idx] = st.session_state.tasks[idx], st.session_state.tasks[idx + 1]

def remove_task(idx: int):
    if 0 <= idx < len(st.session_state.tasks):
        st.session_state.tasks.pop(idx)

def copy_task(idx: int):
    if 0 <= idx < len(st.session_state.tasks):
        row = st.session_state.tasks[idx].copy()
        row["id"] = f"task_copy_{uuid.uuid4().hex[:6]}"
        st.session_state.tasks.insert(idx + 1, row)

def parse_batch_tasks(text: str):
    """
    批次輸入格式：
    任務名稱 ActionBy 工作天數 [0.5天格內文字] [上線日]
    例如：提供素材 Ad2 2天
         客戶確認 客戶 0.5天 1300
         廣告上線 Ad2 1天 上線

    可用單獨一行 -- 作為階段分隔線，系統會在上一筆任務列加入粗下框線。

    也支援省略 Action By：
    任務名稱 工作天數 [上線日]
    省略時會自動判斷：任務名稱含「客戶」則為客戶，其餘預設為 Ad2。
    """
    parsed_rows = []
    errors = []
    launch_words = {"上線", "上線日", "launch", "true", "yes", "y"}

    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue

        if line in {"--", "－－", "—", "——"}:
            if parsed_rows:
                parsed_rows[-1]["粗下框線"] = True
            else:
                errors.append(f"第 {line_no} 行分隔線前沒有可套用的任務：{raw_line}")
            continue

        parts = re.split(r"\s+", line)
        is_launch = False

        if parts and parts[-1].lower() in launch_words:
            is_launch = True
            parts = parts[:-1]

        if len(parts) < 2:
            errors.append(f"第 {line_no} 行格式不足：{raw_line}")
            continue

        def is_days_token(token: str) -> bool:
            token = str(token).strip()
            return token in {"半天", "0.5天", ".5天", "0.5", ".5"} or bool(re.fullmatch(r"\d+(?:\.5)?(?:天)?", token))

        half_day_label = DEFAULT_HALF_DAY_LABEL
        half_label_candidate = ""

        # 先判斷「工作天數 + 0.5文字」格式，避免像 1300、1800 這類半天標註
        # 被誤認成工作天數。例：客戶確認 客戶 0.5天 1300
        if len(parts) >= 3 and is_days_token(parts[-2]):
            days_token = parts[-2]
            half_label_candidate = parts[-1].strip()
            info_parts = parts[:-2]
        elif is_days_token(parts[-1]):
            days_token = parts[-1]
            info_parts = parts[:-1]
        else:
            errors.append(f"第 {line_no} 行工作天數需包含數字，例如 2天、0.5天 或 半天：{raw_line}")
            continue

        if len(info_parts) < 1:
            errors.append(f"第 {line_no} 行缺少任務名稱：{raw_line}")
            continue

        owner_candidate = OWNER_ALIASES.get(info_parts[-1], info_parts[-1]) if len(info_parts) >= 2 else ""
        if len(info_parts) >= 2 and owner_candidate in OWNER_OPTIONS:
            owner = owner_candidate
            task_name = " ".join(info_parts[:-1]).strip()
        else:
            task_name = " ".join(info_parts).strip()
            owner = normalize_owner("", task_name)

        if days_token in {"半天", "0.5天", ".5天", "0.5", ".5"}:
            days = 0.5
        else:
            days_match = re.search(r"\d+(?:\.5)?", days_token)
            if not days_match:
                errors.append(f"第 {line_no} 行工作天數需包含數字，例如 2天、0.5天 或 半天：{raw_line}")
                continue
            days = float(days_match.group())

        if days < 0.5:
            errors.append(f"第 {line_no} 行工作天數需至少 0.5 天：{raw_line}")
            continue
        if abs(days * 2 - round(days * 2)) > 1e-9:
            errors.append(f"第 {line_no} 行工作天數需以 0.5 天為單位，例如 0.5、1、1.5：{raw_line}")
            continue

        if half_label_candidate and abs(days % 1 - 0.5) < 1e-9:
            half_day_label = half_label_candidate

        if not task_name:
            errors.append(f"第 {line_no} 行缺少任務名稱：{raw_line}")
            continue

        parsed_rows.append({
            "id": f"task_batch_{uuid.uuid4().hex[:6]}",
            "顯示": True,
            "任務名稱": task_name,
            "Action By": owner,
            "工作天數": days,
            "半天標註": half_day_label,
            "上線日": is_launch,
            "粗下框線": False,
        })

    if not parsed_rows:
        errors.append("尚未解析到任何可新增的任務。")

    return parsed_rows, errors

def _excel_rgb(cell):
    """回傳儲存格填色 RGB；若沒有可辨識填色則回傳空字串。"""
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type != "solid":
        return ""

    color = fill.fgColor
    if not color:
        return ""

    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()[-6:]

    # 部分 Excel 可能會以 indexed 或 theme 色碼儲存；theme 無法穩定換算 RGB，
    # 但只要它是實心填色，就代表使用者手動標了時程色條，因此保留一個可計算的代碼。
    if color.type == "indexed" and color.indexed is not None:
        indexed_map = {
            64: "",
            22: "D9D9D9",
            10: "FF0000",
            43: "92D050",
        }
        return indexed_map.get(color.indexed, f"INDEXED:{color.indexed}")

    if color.type == "theme" and color.theme is not None:
        tint = getattr(color, "tint", 0) or 0
        return f"THEME:{color.theme}:{round(float(tint), 4)}"

    return ""


def parse_generated_timeline_excel(uploaded_file):
    """
    讀取本工具產出的「時程表」Excel，轉回批次輸入文字。
    判斷邏輯：
    - A 欄：任務名稱
    - B 欄：Action By
    - 第 5 列起：任務列
    - 色條格數：回推工作天數；若格內有半天標註文字，視為半天
    - 紅色色條：標記為上線
    - 其他使用者手動標記的實心色條：也會計入工作天數
    - 「預備上線」列：自動排程產物，不匯入批次流程
    """
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        raise ValueError(f"無法讀取 Excel 檔案，請確認檔案格式是否為 .xlsx：{e}")

    if "時程表" not in workbook.sheetnames:
        raise ValueError("找不到「時程表」工作表，請上傳由此工具產出的時程表 Excel。")

    ws = workbook["時程表"]

    launch_color = EXCEL_COLOR_LAUNCH_BAR.replace("#", "").upper()
    ignored_colors = {
        EXCEL_COLOR_WEEKEND.replace("#", "").upper(),
        "FFFFFF",
        "000000",
    }

    imported_lines = []
    errors = []

    # 工具產出的任務從第 5 列開始；抓到空白任務列即停止。
    for row_idx in range(5, ws.max_row + 1):
        task_name = str(ws.cell(row=row_idx, column=1).value or "").strip()
        owner = str(ws.cell(row=row_idx, column=2).value or "").strip()

        if not task_name:
            continue

        row_colors = [_excel_rgb(ws.cell(row=row_idx, column=col_idx)) for col_idx in range(3, ws.max_column + 1)]

        # 「預備上線」是雙日期模式自動產生的緩衝列，不應回寫成流程項目；
        # 但其他任務即使使用綠色或自訂色，也應計入天數。
        if task_name == "預備上線":
            continue

        bar_days = 0.0
        half_label_from_sheet = ""
        for col_offset, color in enumerate(row_colors, start=3):
            if color and color not in ignored_colors:
                cell_value = str(ws.cell(row=row_idx, column=col_offset).value or "").strip()
                if cell_value:
                    bar_days += 0.5
                    if not half_label_from_sheet:
                        half_label_from_sheet = cell_value
                else:
                    bar_days += 1.0
        is_launch = launch_color in row_colors or "上線" in task_name

        if bar_days <= 0:
            errors.append(f"「{task_name}」沒有讀到可辨識的時程色條，已略過。")
            continue

        owner = normalize_owner(owner, task_name)

        line = f"{task_name} {owner} {format_day_value(bar_days)}天"
        if half_label_from_sheet and abs(bar_days % 1 - 0.5) < 1e-9:
            line += f" {half_label_from_sheet}"
        if is_launch:
            line += " 上線"
        imported_lines.append(line)

        # 若匯入的時程表已有粗下框線，也回寫成批次輸入的 --，方便再次產出時保留分段。
        row_has_thick_bottom = any(
            getattr(ws.cell(row=row_idx, column=col_idx).border.bottom, "style", None) in {"medium", "thick", "double"}
            for col_idx in range(1, ws.max_column + 1)
        )
        if row_has_thick_bottom:
            imported_lines.append("--")

    if not imported_lines:
        extra = "\n" + "\n".join(errors) if errors else ""
        raise ValueError("沒有讀到可匯入的流程項目，請確認檔案是否為此工具匯出的時程表。" + extra)

    return "\n".join(imported_lines), errors


def import_timeline_to_batch(uploaded_file):
    if uploaded_file is None:
        st.session_state.import_msg = "請先選擇要匯入的時程表 Excel。"
        return

    try:
        batch_text, warnings = parse_generated_timeline_excel(uploaded_file)
        st.session_state.batch_tasks_text = batch_text
        warning_text = "\n" + "\n".join(warnings) if warnings else ""
        st.session_state.import_msg = f"已匯入 {len(batch_text.splitlines())} 筆流程到批次輸入。{warning_text}"
        st.session_state.batch_msg = ""
    except Exception as e:
        st.session_state.import_msg = str(e)

def import_timeline_and_apply(uploaded_file):
    """匯入已產出的時程表，並立即套用到流程設定。"""
    if uploaded_file is None:
        st.session_state.import_msg = "請先選擇要匯入的時程表 Excel。"
        return

    try:
        batch_text, warnings = parse_generated_timeline_excel(uploaded_file)
        st.session_state.batch_tasks_text = batch_text

        parsed_rows, errors = parse_batch_tasks(batch_text)
        if errors:
            st.session_state.import_msg = "匯入後解析批次內容時發生錯誤：\n" + "\n".join(errors)
            return

        if sum(1 for row in parsed_rows if row.get("上線日")) > 1:
            st.session_state.import_msg = "匯入內容中只能有一筆標記為上線日。"
            return

        st.session_state.tasks = parsed_rows

        # 清掉舊任務 widget key，避免 Streamlit session state 殘留讓欄位看起來沒更新
        for key in list(st.session_state.keys()):
            if key.startswith(("show_", "task_", "owner_", "days_", "half_label_", "launch_", "up_", "down_", "copy_", "del_")):
                del st.session_state[key]

        warning_text = "\n" + "\n".join(warnings) if warnings else ""
        st.session_state.import_msg = f"已匯入並套用 {len(parsed_rows)} 筆流程。{warning_text}"
        st.session_state.batch_msg = ""
    except Exception as e:
        st.session_state.import_msg = str(e)

def apply_batch_tasks(mode: str = "replace"):
    parsed_rows, errors = parse_batch_tasks(st.session_state.batch_tasks_text)
    if errors:
        st.session_state.batch_msg = "\n".join(errors)
        return

    if sum(1 for row in parsed_rows if row.get("上線日")) > 1:
        st.session_state.batch_msg = "批次輸入中只能有一筆標記為上線日。"
        return

    if mode == "append":
        st.session_state.tasks.extend(parsed_rows)
        action_text = "新增"
    else:
        st.session_state.tasks = parsed_rows
        action_text = "取代"

    # 清掉舊任務 widget key，避免 Streamlit session state 殘留讓欄位看起來沒更新
    for key in list(st.session_state.keys()):
        if key.startswith(("show_", "task_", "owner_", "days_", "half_label_", "launch_", "up_", "down_", "copy_", "del_")):
            del st.session_state[key]

    st.session_state.batch_msg = f"已{action_text} {len(parsed_rows)} 筆任務。"

def load_batch_template():
    # 只載入批次範本文字；保留使用者已設定的開始日期、上線日期與排程方式。
    preserved_settings = {
        key: st.session_state.get(key)
        for key in PROJECT_SETTING_KEYS
        if key in st.session_state
    }

    template_name = st.session_state.get("batch_template_display", BATCH_TEMPLATE_OPTIONS[0])
    st.session_state.batch_tasks_text = BATCH_TEMPLATE_MAP.get(template_name, DEFAULT_BATCH_TASKS_TEXT)
    st.session_state.batch_msg = f"已載入「{template_name}」。"

    for key, value in preserved_settings.items():
        if value is not None:
            st.session_state[key] = value

def generate_schedule():
    had_previous_output = st.session_state.schedule_df is not None

    try:
        holidays = parse_holidays(st.session_state.holidays_text)
        tasks = get_active_tasks()
        calculation_mode = MODE_MAP[st.session_state.mode_display]
        start_date_obj = None if st.session_state.mode_display == "上線日回推" else st.session_state.start_date_value
        launch_date_obj = None if st.session_state.mode_display == "製作日推進" else st.session_state.launch_date_value

        df_schedule, warning_msg, holidays_dt = build_scheduler(
            tasks_config=tasks,
            holidays_config=holidays,
            calculation_mode=calculation_mode,
            start_date_obj=start_date_obj,
            launch_date_obj=launch_date_obj,
        )
        excel_bytes, display_columns = build_excel_bytes(
            df_schedule=df_schedule,
            holidays_config=holidays,
            holidays_dt=holidays_dt,
            launch_date_obj=launch_date_obj,
            collapse_threshold=int(st.session_state.collapse_threshold),
        )
    except Exception as e:
        # 不讓 Streamlit 進入錯誤頁；保留使用者目前輸入與既有預覽。
        st.session_state.validation_error_msg = f"無法產出時程表：{e}"
        st.session_state.status_msg = ""
        return

    st.session_state.schedule_df = df_schedule
    st.session_state.warning_msg = warning_msg
    st.session_state.excel_bytes = excel_bytes
    st.session_state.display_columns = display_columns
    st.session_state.holidays_dt = holidays_dt
    st.session_state.last_generated_name = st.session_state.project_name or "未命名專案"
    st.session_state.validation_error_msg = ""
    st.session_state.status_msg = "時程表已更新。" if had_previous_output else "已產出時程表。"

def reset_defaults():
    st.session_state.project_name = ""
    st.session_state.mode_display = MODE_OPTIONS[0]
    st.session_state.start_date_value = date.today()
    st.session_state.launch_date_value = date.today()
    st.session_state.collapse_threshold = 2
    st.session_state.tasks = [row.copy() for row in DEFAULT_TASKS]
    st.session_state.schedule_df = None
    st.session_state.excel_bytes = None
    st.session_state.display_columns = None
    st.session_state.holidays_dt = None
    st.session_state.warning_msg = ""
    st.session_state.last_generated_name = "未命名專案"
    st.session_state.status_msg = ""
    st.session_state.batch_tasks_text = DEFAULT_BATCH_TASKS_TEXT
    st.session_state.batch_template_display = BATCH_TEMPLATE_OPTIONS[0]
    st.session_state.batch_msg = ""
    st.session_state.import_msg = ""
    st.session_state.half_day_label = "1300"

# =========================
# UI
# =========================
st.title("製作時程排程工具")
st.caption("快速設定專案日期與流程後，即可產出 Excel 時程表。")

with st.sidebar:
    st.subheader("假日設定")
    st.text_area("假日清單（每行一筆，格式：YYYY-MM-DD,名稱）", key="holidays_text", height=420)

with st.container(border=True):
    c1, c2 = st.columns([6,0.65], vertical_alignment="center")
    with c1:
        st.markdown('<div class="section-title">專案設定</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">先決定排程方式與日期，再按下「產出時程表」。</div>', unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="mini-reset">', unsafe_allow_html=True)
        st.button("重設", on_click=reset_defaults, use_container_width=True)

    r1c1, r1c2, r1c3 = st.columns([2.6,1.5,1.0], vertical_alignment="bottom")
    with r1c1:
        st.text_input("專案名稱", key="project_name", placeholder="請輸入專案名稱")
    with r1c2:
        st.selectbox("排程方式", MODE_OPTIONS, key="mode_display")
    with r1c3:
        st.number_input("日期縮略門檻", min_value=1, max_value=30, step=1, key="collapse_threshold")

    start_disabled = st.session_state.mode_display == "上線日回推"
    launch_disabled = st.session_state.mode_display == "製作日推進"

    r2c1, r2c2, r2c3 = st.columns([1.35,1.35,1.0], vertical_alignment="bottom")
    with r2c1:
        st.date_input("開始日期", key="start_date_value", disabled=start_disabled)
    with r2c2:
        st.date_input("上線日期", key="launch_date_value", disabled=launch_disabled)
    with r2c3:
        st.button("產出時程表", type="primary", use_container_width=True, on_click=generate_schedule)

st.markdown('<div class="small-gap"></div>', unsafe_allow_html=True)

if st.session_state.validation_error_msg:
    st.error(st.session_state.validation_error_msg)

if st.session_state.warning_msg:
    st.warning(st.session_state.warning_msg)

if st.session_state.status_msg:
    st.success(st.session_state.status_msg)

if st.session_state.schedule_df is not None:
    with st.container(border=True):
        p1, p2 = st.columns([5.4,1.05], vertical_alignment="center")
        with p1:
            st.markdown('<div class="section-title">排程預覽</div>', unsafe_allow_html=True)
        with p2:
            filename = f"{datetime.now().strftime('%m%d')}_{st.session_state.last_generated_name}.xlsx"
            st.download_button("下載 Excel", data=st.session_state.excel_bytes, file_name=filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, type="primary")
        st.markdown(
            render_stable_preview(
                st.session_state.schedule_df,
                st.session_state.display_columns,
                st.session_state.holidays_dt,
            ),
            unsafe_allow_html=True,
        )

st.markdown('<div class="small-gap"></div>', unsafe_allow_html=True)




manual_tab, batch_tab = st.tabs(["單筆編輯", "批次輸入"])

with manual_tab:
    with st.container(border=True):
        h1, h2 = st.columns([5,1.05], vertical_alignment="center")
        with h1:
            st.markdown('<div class="section-title">流程設定</div>', unsafe_allow_html=True)
            st.markdown('<div class="section-sub">可新增、複製、刪除、排序與修改任務。</div>', unsafe_allow_html=True)
        with h2:
            st.button("新增任務", on_click=add_task, use_container_width=True)

        hc1, hc2, hc3, hc4, hc5, hc6, hc7, hc8 = st.columns([0.55, 2.85, 1.15, 1.85, 0.68, 1.05, 0.55, 0.55], vertical_alignment="center")
        headers = [
            (hc1, "顯示"),
            (hc2, "任務名稱"),
            (hc3, "Action By"),
            (hc4, "工作天數"),
            (hc5, "上線日"),
            (hc6, "排序"),
            (hc7, "複製"),
            (hc8, "刪除"),
        ]
        centered_headers = {"顯示", "工作天數", "上線日", "排序", "複製", "刪除"}
        for col, label in headers:
            with col:
                cls = "task-head-label task-head-center" if label in centered_headers else "task-head-label"
                st.markdown(f'<div class="{cls}">{label}</div>', unsafe_allow_html=True)

        for idx, row in enumerate(st.session_state.tasks):
            rid = row["id"]
            c1, c2, c3, c4, c5, c6, c7, c8 = st.columns([0.55, 2.85, 1.15, 1.85, 0.68, 1.05, 0.55, 0.55], vertical_alignment="center")

            with c1:
                key = f"show_{rid}"
                if key not in st.session_state:
                    st.session_state[key] = row["顯示"]
                cc1, cc2, cc3 = st.columns([1, 0.9, 1], vertical_alignment="center")
                with cc2:
                    st.checkbox("顯示", key=key, label_visibility="collapsed",
                                on_change=sync_task_field, args=(rid, "顯示", key))

            with c2:
                key = f"task_{rid}"
                if key not in st.session_state:
                    st.session_state[key] = row["任務名稱"]
                st.text_input("任務名稱", key=key, label_visibility="collapsed",
                              on_change=sync_task_field, args=(rid, "任務名稱", key))

            with c3:
                key = f"owner_{rid}"
                if key not in st.session_state:
                    st.session_state[key] = row["Action By"]
                st.selectbox("Action By", OWNER_OPTIONS, key=key, label_visibility="collapsed",
                             on_change=sync_task_field, args=(rid, "Action By", key))

            with c4:
                days_key = f"days_{rid}"
                if days_key not in st.session_state:
                    st.session_state[days_key] = float(row["工作天數"])
                current_days_value = float(st.session_state.get(days_key, row.get("工作天數", 1.0)) or 1.0)
                has_half_day = abs((current_days_value * 2) % 2 - 1) < 1e-9
                half_key = f"half_label_{rid}"
                if half_key not in st.session_state:
                    st.session_state[half_key] = row.get("半天標註", DEFAULT_HALF_DAY_LABEL)

                if has_half_day:
                    dc1, dc2 = st.columns([0.9, 1.1], vertical_alignment="center")
                    with dc1:
                        st.number_input("工作天數", min_value=0.5, step=0.5, format="%.1f", key=days_key, label_visibility="collapsed",
                                        on_change=sync_task_field, args=(rid, "工作天數", days_key))
                    with dc2:
                        st.text_input("0.5文字", key=half_key, label_visibility="collapsed", placeholder="1300",
                                      on_change=sync_task_field, args=(rid, "半天標註", half_key))
                else:
                    st.number_input("工作天數", min_value=0.5, step=0.5, format="%.1f", key=days_key, label_visibility="collapsed",
                                    on_change=sync_task_field, args=(rid, "工作天數", days_key))

            with c5:
                key = f"launch_{rid}"
                if key not in st.session_state:
                    st.session_state[key] = row["上線日"]
                lc1, lc2, lc3 = st.columns([1, 0.9, 1], vertical_alignment="center")
                with lc2:
                    st.checkbox("上線日", key=key, label_visibility="collapsed",
                                on_change=sync_task_field, args=(rid, "上線日", key))

            with c6:
                s1, s2 = st.columns([1, 1], vertical_alignment="center")
                with s1:
                    if st.button("↑", key=f"up_{rid}", use_container_width=True, disabled=(idx == 0)):
                        move_task_up(idx)
                        st.rerun()
                with s2:
                    if st.button("↓", key=f"down_{rid}", use_container_width=True, disabled=(idx == len(st.session_state.tasks) - 1)):
                        move_task_down(idx)
                        st.rerun()

            with c7:
                if st.button("⧉", key=f"copy_{rid}", use_container_width=True):
                    copy_task(idx)
                    st.rerun()

            with c8:
                if st.button("✕", key=f"del_{rid}", use_container_width=True):
                    remove_task(idx)
                    st.rerun()

            if idx < len(st.session_state.tasks) - 1:
                st.markdown('<div class="flow-row-sep"></div>', unsafe_allow_html=True)

with batch_tab:
    with st.container(border=True):
        st.markdown('<div class="section-title">多時程項目批次輸入</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="section-sub">每行一筆任務，使用空白區隔資訊。格式：任務名稱 ActionBy 工作天數 [0.5文字] [上線]；工作天數支援 0.5天／半天，例如「客戶確認 客戶 0.5天 1300」。若輸入單獨一行 --，會在上一筆任務下方加入粗分隔線。</div>',
            unsafe_allow_html=True,
        )

        uploaded_timeline_file = st.file_uploader(
            "匯入已產出的時程表",
            type=["xlsx"],
            help="上傳由此工具下載的時程表 Excel，可自動將任務名稱、Action By 與工作天數帶回批次輸入；若有手動改色的特殊需求色條，也會納入天數計算。",
        )
        import_col1, import_col2 = st.columns([1.35, 4.65], vertical_alignment="center")
        with import_col1:
            if st.button("匯入並套用", use_container_width=True):
                import_timeline_and_apply(uploaded_timeline_file)
                st.rerun()
        with import_col2:
            st.caption("上傳後會自動寫入批次輸入區，並直接取代目前流程。")

        if st.session_state.import_msg:
            if "無法" in st.session_state.import_msg or "找不到" in st.session_state.import_msg or "沒有讀到" in st.session_state.import_msg or "請先" in st.session_state.import_msg:
                st.warning(st.session_state.import_msg)
            else:
                st.success(st.session_state.import_msg)

        t1, t2 = st.columns([2, 1], vertical_alignment="bottom")
        with t1:
            st.selectbox("套用範本", BATCH_TEMPLATE_OPTIONS, key="batch_template_display")
        with t2:
            if st.button("載入範本", use_container_width=True):
                load_batch_template()
                st.rerun()

        st.text_area(
            "批次輸入內容",
            key="batch_tasks_text",
            height=280,
            placeholder="提供素材 Ad2 2天\n客戶確認 客戶 1天\n廣告上線 Ad2 1天 上線",
        )
        b1, b2 = st.columns([1, 5], vertical_alignment="center")
        with b1:
            if st.button("套用到流程", use_container_width=True):
                apply_batch_tasks("replace")
                st.rerun()
        with b2:
            st.caption("套用後會取代目前流程；若沒有標記「上線」，系統仍會沿用原本邏輯：產出時自動將最後一筆視為上線日。")

        if st.session_state.batch_msg:
            if "錯誤" in st.session_state.batch_msg or "第 " in st.session_state.batch_msg or "只能" in st.session_state.batch_msg or "尚未" in st.session_state.batch_msg:
                st.warning(st.session_state.batch_msg)
            else:
                st.success(st.session_state.batch_msg)
